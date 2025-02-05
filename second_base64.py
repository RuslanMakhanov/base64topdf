import os
import csv
import re
import base64
from openpyxl import load_workbook
from datetime import datetime
import sys

# Логи сохраняются в текстовые файлы
ACCESS_LOG_FILE = "./logs/access_log.txt"
ERROR_LOG_FILE = "./logs/error_log.txt"
CHECK_LOG_FILE = "./logs/check.txt"

def log_message(message, log_type="access"):
    if log_type == "access":
        log_file = ACCESS_LOG_FILE
    elif log_type == "error":
        log_file = ERROR_LOG_FILE
    elif log_type == "check":
        log_file = CHECK_LOG_FILE
    else:
        raise ValueError(f"Неизвестный тип лога: {log_type}")
    
    with open(log_file, "a", encoding="utf-8") as log:
        log.write(message + "\n")


csv.field_size_limit(sys.maxsize) # Устанавливаем разумное ограничение

def process_csv_and_excel(csv_file, excel_file, output_dir):
    # Проверяем, существует ли CSV файл
    if not os.path.exists(csv_file):
        print(f"CSV файл не найден: {csv_file}")
        log_message(f"CSV файл не найден: {csv_file}", log_type="error")
        return

    # Проверяем, существует ли Excel файл
    if not os.path.exists(excel_file):
        print(f"Excel файл не найден: {excel_file}")
        log_message(f"Excel файл не найден: {excel_file}", log_type="error")
        return

    # Открываем Excel файл
    try:
        wb = load_workbook(excel_file)
        sheet = wb.active
    except Exception as e:
        print(f"Ошибка при открытии Excel файла: {e}")
        log_message(f"Ошибка при открытии Excel файла: {e}", log_type="error")
        return

    # Проверяем и создаем папку для сохранения файлов
    if not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)

    # Регулярные выражения для извлечения данных
    unique_code_pattern = re.compile(r'101000\d+')
    base64_pattern = re.compile(r'data:application/pdf;base64,[A-Za-z0-9+/=]+')

    # Читаем CSV файл построчно
    try:
        with open(csv_file, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            for row in reader:
                content = ",".join(row) # Собираем строку обратно

                # Извлекаем уникальный код
                unique_code_match = unique_code_pattern.search(content)
                if unique_code_match:
                    unique_code = unique_code_match.group()
                    print(f"Уникальный код извлечен: {unique_code}")
                    log_message(f"\nУникальный код извлечен: {unique_code}")
                else:
                    print("Уникальный код не найден в строке CSV")
                    log_message("Уникальный код не найден в строке CSV", log_type="error")
                    continue

                # Извлекаем Base64 данные
                base64_matches = base64_pattern.findall(content)
                base64_match = base64_pattern.search(content)
                if base64_match:
                    base64_data = base64_match.group().split(",")[1] # Убираем префикс "dat a:application/pdf;base64,"
                    print("Base64 данные успешно извлечены")
                    log_message("Base64 данные успешно извлечены")
                    if len(base64_matches) > 1: # Проверяем, есть ли второй Base64
                        base64_data_second = base64_matches[1].split(",")[1] # Второй Base64
                        print("Второй Base64 данные успешно извлечены")
                        log_message("Второй Base64"+ unique_code +" данные успешно извлечены" + base64_data_second, log_type="check")
                    else:
                        print("Второй Base64 код не найден в строке CSV")
                        log_message("Второй Base64 код не найден в строке CSV", log_type="error")

                else:
                    print("Base64 данные не найдены в строке CSV")
                    log_message("Base64 данные не найдены в строке CSV", log_type="error")
                    continue

                # Ищем строку в Excel по уникальному коду и извлекаем название файла и дату
                file_name = None
                date_str = None
                for row in sheet.iter_rows(values_only=True):
                    excel_code = str(row[0]).strip() # Удаляем лишние пробелы
                    if unique_code == excel_code: # Сравниваем уникальный код
                        print(f"Совпадение найдено: {unique_code} == {excel_code}")
                        log_message(f"Совпадение найдено: {unique_code} == {excel_code}")

                        # Извлекаем имя файла из столбца D
                        if row[3]:
                            if "{{" in row[4] and "}" in row[4]:
                                try:
                                    file_name = row[4].split("{{")[1].split("}")[0]
                                    print(f"Имя файла: {file_name}")
                                    log_message(f"Имя файла: {file_name}")
                                except IndexError:
                                    print(f"Ошибка извлечения имени файла для УК {unique_code}: {row[4]}")
                                    log_message(f"Ошибка извлечения имени файла для УК {unique_code}: {row[4]}", log_type="error")
                        else:
                            print(f"Столбец D пуст для УК {unique_code}")
                            log_message(f"Столбец D пуст для УК {unique_code}", log_type="error")

                        # Извлекаем дату из столбца B
                        if row[1]:
                            try:
                                date_str = row[1].strip()
                                date_obj = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
                                folder_name = date_obj.strftime("%m%d") # Формат "месяц день"
                                print(f"Дата: {folder_name}")
                            except Exception as e:
                                print(f"Ошибка преобразования даты для УК {unique_code}: {e}")
                                log_message(f"Ошибка преобразования даты для УК {unique_code}: {e}", log_type="error")
                                continue
                        else:
                            print(f"Столбец B пуст для УК {unique_code}")
                            log_message(f"Столбец B пуст для УК {unique_code}", log_type="error")
                        break

                if not file_name or not date_str:
                    print(f"Данные для УК {unique_code} не найдены в Excel.")
                    log_message(f"Данные для УК {unique_code} не найдены в Excel.", log_type="error")
                    continue

                # Создаем папку с месяцем и днем
                date_folder = os.path.join(output_dir, folder_name)
                os.makedirs(date_folder, exist_ok=True)

                # Полный путь к файлу
                file_path = os.path.join(date_folder, file_name)

                # Создаем файл и вставляем Base64 код
                try:
                    with open(file_path, "wb") as f:
                        f.write(base64.b64decode(base64_data_second))
                    print(f"Файл сохранен: {file_path}")
                    log_message(f"Файл сохранен: {file_path}")
                except Exception as e:
                    print(f"Ошибка при сохранении файла {file_path}: {e}")
                    log_message(f"Ошибка при сохранении файла {file_path}: {e}", log_type="error")
    except Exception as e:
        print(f"Ошибка при обработке CSV файла: {e}")
        log_message(f"Ошибка при обработке CSV файла: {e} \n", log_type="error")

# Пример вызова функции
csv_path = "./Apostil_SEP.csv" # Путь к CSV файлуprint(f"Уникальный код извлечен: {unique_code}")
excel_path = "./excel/apostil_september_v2-org.xlsx" # Путь к Excel файлу
output_directory = "./saves" # Папка для сохранения файлов

process_csv_and_excel(csv_path, excel_path, output_directory)